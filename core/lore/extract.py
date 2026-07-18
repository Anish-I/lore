"""Text extraction for non-markdown sources (M4): PDF, DOCX, and XLSX.

PDF via PyMuPDF (fitz) — already a dependency of the local stack.
DOCX via stdlib zipfile + XML (a .docx is a zip; word/document.xml holds the
runs) — no python-docx dependency needed.
XLSX via openpyxl (read-only stream) — spreadsheets are core "people-work"
content for the enterprise tool; cells render as pipe-delimited rows so the
chunker keeps header<->value structure.

extract_text(path) -> (title, text) or None when the format is unsupported.
Output is markdown-ish plain text with a leading H1 so the chunker has
structure to work with.
"""
import os
import re
import zipfile
from xml.etree import ElementTree

EXTRACTABLE_EXTS = {".pdf", ".docx", ".xlsx"}

# Bound a decompression-bomb spreadsheet: openpyxl read-only streams rows so
# memory stays flat, but cap total cells anyway so a crafted sheet can't spin.
_XLSX_MAX_CELLS = 200_000

_W_NS = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"

# Guard against a malicious .docx: cap the uncompressed size of the document
# part so a zip-bomb can't exhaust memory during read().
_DOCX_MAX_UNCOMPRESSED = 64 * 1024 * 1024  # 64 MB of XML is already absurd


_DTD_RE = re.compile(rb"<!DOCTYPE|<!ENTITY", re.IGNORECASE)


def _safe_fromstring(xml_bytes):
    """Parse .docx XML with entity/DTD attacks defused.

    A .docx document part never legitimately carries a DTD or custom entities,
    so reject any that does BEFORE parsing — this closes billion-laughs (internal
    entity expansion) and XXE (external entities) without a third-party parser.
    """
    if _DTD_RE.search(xml_bytes[:4096]):
        raise ValueError("DTD/entity declaration not allowed in .docx XML")
    return ElementTree.fromstring(xml_bytes)


def _pdf_text(path: str) -> str:
    import fitz  # PyMuPDF
    parts = []
    with fitz.open(path) as doc:
        for page in doc:
            t = page.get_text("text")
            if t and t.strip():
                parts.append(t.strip())
    return "\n\n".join(parts)


def _docx_text(path: str) -> str:
    with zipfile.ZipFile(path) as z:
        info = z.getinfo("word/document.xml")
        if info.file_size > _DOCX_MAX_UNCOMPRESSED:
            raise ValueError(f"docx document part too large ({info.file_size} bytes)")
        xml = z.read("word/document.xml")
    root = _safe_fromstring(xml)
    paras = []
    for p in root.iter(f"{_W_NS}p"):
        runs = [t.text or "" for t in p.iter(f"{_W_NS}t")]
        line = "".join(runs).strip()
        if line:
            paras.append(line)
    return "\n\n".join(paras)


def _xlsx_text(path: str) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        blocks = []
        cells = 0
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows(values_only=True):
                vals = ["" if v is None else str(v) for v in row]
                cells += len(vals)
                if cells > _XLSX_MAX_CELLS:
                    raise ValueError(f"xlsx too large (> {_XLSX_MAX_CELLS} cells)")
                if any(v.strip() for v in vals):
                    rows.append(" | ".join(vals).rstrip(" |"))
            if rows:
                blocks.append(f"## {ws.title}\n" + "\n".join(rows))
        return "\n\n".join(blocks)
    finally:
        wb.close()


def extract_text(path: str):
    """Return (title, markdown_text) for supported binary formats, else None.

    A hostile or malformed document (DTD bomb, zip bomb, corrupt PDF) is treated
    as unextractable → None; never raises to the caller."""
    ext = os.path.splitext(path)[1].lower()
    if ext not in EXTRACTABLE_EXTS:
        return None
    try:
        if ext == ".pdf":
            body = _pdf_text(path)
        elif ext == ".docx":
            body = _docx_text(path)
        else:
            body = _xlsx_text(path)
    except Exception:
        return None
    body = re.sub(r"\n{3,}", "\n\n", body or "").strip()
    if not body:
        return None
    title = os.path.splitext(os.path.basename(path))[0]
    return title, f"# {title}\n\n{body}\n"
